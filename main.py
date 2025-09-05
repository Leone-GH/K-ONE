# main.py
# pip install playwright openpyxl python-dotenv keyring
# playwright install chromium

from pathlib import Path
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
from dotenv import load_dotenv
import keyring, getpass
import argparse
import os, re, sys

EXCEL_PATH = Path("2025 8월 자동입력 차량일지양식.xlsx")
SHEET_NAME = None
START_ROW  = 9
TARGET_URL = os.getenv("NKP_URL", "http://intra.k-one.co.kr/car_drive_add.asp")

LOGIN_URL_HINTS = ["login", "member/login", "car_login"]

LOGIN_SELECTORS = {
    "id":  'input[name="user_id"], #user_id, #userid, input#id',
    "pw":  'input[name="user_pw"], #user_pw, #userpw, input[type="password"]',
    "btn": 'button[type="submit"], input[type="submit"], #loginBtn',
    "ok_any": ['text=로그아웃', 'a[href*="logout"]', 'text=메인', 'nav .user'],
    "fail_any": ['.alert-danger', '.error', 'text=아이디', 'text=비밀번호']
}

POST_SUBMIT_OK_SELECTORS = [
    ".toast.success", ".alert-success",
    "text=등록되었습니다", "text=정상적으로 처리"
]

SERVICE_NAME = "NKP_AUTOFILL"

def thousands_sep(x: str) -> str:
    s = re.sub(r"[^\d.]", "", str(x or ""))
    if not s:
        return ""
    if "." in s:
        n, frac = s.split(".", 1)
        return f"{int(n):,}.{frac}"
    return f"{int(s):,}"

def hhmm_parts(s: str):
    if not s:
        return ("", "")
    t = str(s).strip().replace(":", "")
    if len(t) < 3:
        return ("", "")
    return (t[:-2].zfill(2), t[-2:].zfill(2))

def norm_company(name: str) -> str:
    if not name: return ""
    n = str(name).strip()
    table = {
        "광주지사": "본사(회사)", "본사": "본사(회사)",
        "국민연금": "국민연금공단", "건보": "국민건강보험공단",
        "신보": "신용보증기금", "롯데손보": "롯데손해보험", "한화손보": "한화손해보험",
    }
    return table.get(n, n)

def norm_work_kind(s: str) -> str:
    if not s: return ""
    n = str(s).strip()
    table = {"방문처리": "장애처리", "원격처리": "장애처리", "장비회수": "반납(회수)", "협업": "업무협의"}
    return table.get(n, n)

def norm_endpoint(s: str) -> str:
    if not s: return ""
    n = str(s).strip()
    return "광주 서구 상일로 24번길 19" if n == "본사" else n

def wait_any(page, selectors, timeout=6000) -> bool:
    for sel in selectors:
        try:
            page.wait_for_selector(sel, timeout=timeout)
            return True
        except PWTimeout:
            continue
    return False

def is_login_page(page) -> bool:
    url = page.url.lower()
    if any(h in url for h in [h.lower() for h in LOGIN_URL_HINTS]):
        return True
    if page.locator(LOGIN_SELECTORS["id"]).count() and page.locator(LOGIN_SELECTORS["pw"]).count():
        return True
    return False

def login_if_needed(page, args) -> None:
    if "본 페이지에 접근 권한이 없습니다" in page.content():
        page.goto(TARGET_URL, wait_until="domcontentloaded")

    def do_login(user_id: str, user_pw: str):
        if not is_login_page(page):
            page.goto(TARGET_URL, wait_until="domcontentloaded")
        page.locator(LOGIN_SELECTORS["id"]).first.fill(user_id)
        page.locator(LOGIN_SELECTORS["pw"]).first.fill(user_pw)
        if page.locator(LOGIN_SELECTORS["btn"]).count():
            page.locator(LOGIN_SELECTORS["btn"]).first.click()
        else:
            page.evaluate("""() => { const f = document.querySelector('form'); f && f.submit(); }""")
        ok = wait_any(page, LOGIN_SELECTORS["ok_any"], timeout=8000)
        if not ok:
            if wait_any(page, LOGIN_SELECTORS["fail_any"], timeout=2000):
                raise RuntimeError("로그인 실패로 보입니다. ID/PW를 확인하세요.")
            page.wait_for_timeout(1200)

    env_id = os.getenv("NKP_ID")
    env_pw = os.getenv("NKP_PW")
    if env_id and env_pw:
        do_login(env_id, env_pw)
        return

    stored_id = keyring.get_password(SERVICE_NAME, "last_user")
    if stored_id:
        stored_pw = keyring.get_password(SERVICE_NAME, stored_id)
        if stored_pw:
            do_login(stored_id, stored_pw)
            return

    print("[NKP] 저장된 계정이 없습니다. 최초 1회 ID/PW를 입력해 저장합니다.")
    user_id = input("NKP ID: ").strip()
    user_pw = getpass.getpass("NKP PW: ").strip()
    if not user_id or not user_pw:
        raise SystemExit("ID/PW가 비었습니다.")
    keyring.set_password(SERVICE_NAME, "last_user", user_id)
    keyring.set_password(SERVICE_NAME, user_id, user_pw)
    do_login(user_id, user_pw)

def reset_credentials():
    stored_id = keyring.get_password(SERVICE_NAME, "last_user")
    if stored_id:
        try: keyring.delete_password(SERVICE_NAME, "last_user")
        except: pass
        try: keyring.delete_password(SERVICE_NAME, stored_id)
        except: pass
    print("저장된 NKP 자격증명을 삭제했습니다.")

def main():
    load_dotenv()
    parser = argparse.ArgumentParser(description="NKP 차량일지 자동입력")
    parser.add_argument("--reset-creds", action="store_true", help="저장된 ID/PW 삭제")
    args = parser.parse_args()
    if args.reset_creds:
        reset_credentials()
        return

    if not EXCEL_PATH.exists():
        raise SystemExit(f"엑셀 파일이 없습니다: {EXCEL_PATH}")

    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb[SHEET_NAME] if SHEET_NAME else wb[wb.sheetnames[0]]

    with sync_playwright() as p:
        ctx = p.chromium.launch_persistent_context(
            user_data_dir=str(Path("chrome_profile").resolve()),
            headless=False,
            args=["--disable-blink-features=AutomationControlled"],
        )
        page = ctx.new_page()
        page.goto(TARGET_URL, wait_until="domcontentloaded")

        if is_login_page(page) or "본 페이지에 접근 권한이 없습니다" in page.content():
            login_if_needed(page, args)
            page.goto(TARGET_URL, wait_until="domcontentloaded")

        r = START_ROW
        while True:
            date_val  = ws.cell(row=r, column=1).value
            comp_val  = ws.cell(row=r, column=6).value
            if not comp_val:
                print("모든 행 처리 완료")
                break

            stime = ws.cell(row=r, column=17).value
            etime = ws.cell(row=r, column=27).value
            sh, sm = hhmm_parts("" if stime is None else str(stime))
            eh, em = hhmm_parts("" if etime is None else str(etime))

            start_km = ws.cell(row=r, column=4).value
            end_km   = ws.cell(row=r, column=8).value
            far_val  = ws.cell(row=r, column=10).value
            kind_val = ws.cell(row=r, column=11).value
            park_fee = ws.cell(row=r, column=15).value
            toll_fee = ws.cell(row=r, column=16).value
            end_point = ws.cell(row=r, column=7).value

            comp_out = norm_company(comp_val)
            kind_out = norm_work_kind(kind_val)
            end_point_out = norm_endpoint(end_point)

            page.fill("#datepicker", "" if date_val is None else str(date_val))

            if sh: page.select_option("#start_hh", value=sh) if page.locator("#start_hh").count() else page.fill("#start_hh", sh)
            if sm: page.select_option("#start_mm", value=sm) if page.locator("#start_mm").count() else page.fill("#start_mm", sm)
            if eh: page.select_option("#end_hh",   value=eh) if page.locator("#end_hh").count() else page.fill("#end_hh", eh)
            if em: page.select_option("#end_mm",   value=em) if page.locator("#end_mm").count() else page.fill("#end_mm", em)

            page.fill("#start_km", thousands_sep(start_km))
            page.fill("#end_km",   thousands_sep(end_km))
            page.fill("#far",      "" if far_val is None else str(far_val))

            page.locator('[name="end_company"]').first.fill(comp_out)
            page.fill("#end_point", end_point_out)
            page.locator('[name="run_memo"]').first.fill(kind_out)

            page.fill("#parking", "" if park_fee is None else str(park_fee))
            page.fill("#toll",    "" if toll_fee is None else str(toll_fee))

            if page.locator('button[type="submit"]').count():
                page.click('button[type="submit"]')
            elif page.locator('input[type="submit"]').count():
                page.click('input[type="submit"]')
            else:
                page.evaluate("""() => { const f = document.querySelector('form'); f && f.submit(); }""")

            ok = wait_any(page, POST_SUBMIT_OK_SELECTORS, timeout=15000)
            if not ok:
                try: page.keyboard.press("Enter")
                except: pass
                page.screenshot(path=f"row_{r}_error.png", full_page=True)
                print(f"[경고] {r}행: 제출 성공 신호 미감지(스냅샷 저장)")

            r += 1

        ctx.close()

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"[오류] {e}")
        sys.exit(1)
